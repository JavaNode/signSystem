from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import List
from ..database import get_db
from ..services.participant_service import ParticipantService
from ..services.judge_service import JudgeService
from ..services.group_service import GroupService
from ..services.checkin_service import CheckinService
from ..services.score_service import ScoreService
from ..utils.qr_generator import create_qr_code_sheet
import openpyxl
import io

router = APIRouter()

class ImportResult(BaseModel):
    success_count: int
    error_count: int
    errors: List[dict]

@router.post("/import/participants")
async def import_participants_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """从Excel导入参赛者数据"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件")
    
    try:
        # 读取Excel文件
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
        
        participants_data = []
        errors = []
        
        # 跳过标题行，从第2行开始读取
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # 跳过空行
                    continue
                
                name, organization, phone, group_name, photo_filename = row[:5]
                
                if not all([name, organization, phone]):
                    errors.append({
                        "row": row_num,
                        "error": "姓名、单位、手机号不能为空"
                    })
                    continue
                
                # 查找组别ID
                group_id = None
                if group_name:
                    # 这里可以根据组名查找组ID，或者创建新组
                    pass
                
                participants_data.append({
                    "name": str(name).strip(),
                    "organization": str(organization).strip(),
                    "phone": str(phone).strip(),
                    "group_id": group_id,
                    "photo_path": f"data/photos/{photo_filename}" if photo_filename else None
                })
                
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": f"数据解析错误: {str(e)}"
                })
        
        # 批量创建参赛者
        created_participants = []
        if participants_data:
            try:
                created_participants = ParticipantService.batch_create_participants(db, participants_data)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"批量创建失败: {str(e)}")
        
        return ImportResult(
            success_count=len(created_participants),
            error_count=len(errors),
            errors=errors
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件处理失败: {str(e)}")

@router.post("/import/judges")
async def import_judges_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """从Excel导入评委数据"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="请上传Excel文件")
    
    try:
        # 读取Excel文件
        content = await file.read()
        workbook = openpyxl.load_workbook(io.BytesIO(content))
        sheet = workbook.active
        
        judges_data = []
        errors = []
        
        # 跳过标题行，从第2行开始读取
        for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                if not any(row):  # 跳过空行
                    continue
                
                name, username, password, organization = row[:4]
                
                if not all([name, username, password]):
                    errors.append({
                        "row": row_num,
                        "error": "姓名、用户名、密码不能为空"
                    })
                    continue
                
                judges_data.append({
                    "name": str(name).strip(),
                    "username": str(username).strip(),
                    "password": str(password).strip(),
                    "organization": str(organization).strip() if organization else None
                })
                
            except Exception as e:
                errors.append({
                    "row": row_num,
                    "error": f"数据解析错误: {str(e)}"
                })
        
        # 批量创建评委
        created_judges = []
        if judges_data:
            try:
                created_judges = JudgeService.batch_create_judges(db, judges_data)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"批量创建失败: {str(e)}")
        
        return ImportResult(
            success_count=len(created_judges),
            error_count=len(errors),
            errors=errors
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"文件处理失败: {str(e)}")

@router.post("/generate/qr-sheet")
async def generate_qr_code_sheet(db: Session = Depends(get_db)):
    """生成二维码打印表格"""
    try:
        participants = ParticipantService.get_all_participants(db)
        
        if not participants:
            raise HTTPException(status_code=400, detail="没有参赛者数据")
        
        file_path = create_qr_code_sheet(participants)
        
        return {
            "message": "二维码表格生成成功",
            "file_path": file_path,
            "participant_count": len(participants)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"生成失败: {str(e)}")

@router.get("/export/participants")
async def export_participants_excel(db: Session = Depends(get_db)):
    """导出参赛者数据到Excel"""
    try:
        participants = ParticipantService.get_all_participants(db)
        
        # 创建Excel工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "参赛者名单"
        
        # 设置标题行
        headers = ["ID", "姓名", "单位", "手机号", "组别", "是否签到", "签到时间", "平均分", "二维码ID"]
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 填充数据
        for row, participant in enumerate(participants, 2):
            sheet.cell(row=row, column=1, value=participant.id)
            sheet.cell(row=row, column=2, value=participant.name)
            sheet.cell(row=row, column=3, value=participant.organization)
            sheet.cell(row=row, column=4, value=participant.phone)
            sheet.cell(row=row, column=5, value=participant.group.name if participant.group else "未分组")
            sheet.cell(row=row, column=6, value="是" if participant.is_checked_in else "否")
            sheet.cell(row=row, column=7, value=participant.checkin_time.strftime('%Y-%m-%d %H:%M:%S') if participant.checkin_time else "")
            sheet.cell(row=row, column=8, value=participant.average_score)
            sheet.cell(row=row, column=9, value=participant.qr_code_id)
        
        # 保存到内存
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return {
            "message": "导出成功",
            "filename": "participants.xlsx",
            "data": output.getvalue().hex()  # 转换为十六进制字符串传输
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

@router.get("/export/scores")
async def export_scores_excel(round_number: int = 1, db: Session = Depends(get_db)):
    """导出评分数据到Excel"""
    try:
        export_data = ScoreService.export_scores(db, round_number)
        
        # 创建Excel工作簿
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = f"第{round_number}轮评分结果"
        
        if not export_data["data"]:
            raise HTTPException(status_code=400, detail="没有评分数据")
        
        # 获取所有列名
        first_row = export_data["data"][0]
        headers = list(first_row.keys())
        
        # 设置标题行
        for col, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col, value=header)
        
        # 填充数据
        for row, data in enumerate(export_data["data"], 2):
            for col, header in enumerate(headers, 1):
                sheet.cell(row=row, column=col, value=data.get(header, ""))
        
        # 保存到内存
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return {
            "message": "导出成功",
            "filename": f"scores_round_{round_number}.xlsx",
            "data": output.getvalue().hex()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

@router.post("/reset/all-checkins")
async def reset_all_checkins(db: Session = Depends(get_db)):
    """重置所有签到状态（危险操作）"""
    try:
        participants = ParticipantService.get_all_participants(db)
        
        reset_count = 0
        for participant in participants:
            if participant.is_checked_in:
                ParticipantService.update_participant(db, participant.id, {
                    "is_checked_in": False,
                    "checkin_time": None
                })
                reset_count += 1
        
        return {
            "message": f"已重置 {reset_count} 个参赛者的签到状态",
            "reset_count": reset_count
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"重置失败: {str(e)}")

@router.get("/system/status")
async def get_system_status(db: Session = Depends(get_db)):
    """获取系统状态"""
    try:
        # 获取各种统计数据
        participant_stats = ParticipantService.get_participants_statistics(db)
        group_stats = GroupService.get_groups_statistics(db)
        judge_stats = JudgeService.get_judge_statistics(db)
        
        return {
            "participants": participant_stats,
            "groups": group_stats,
            "judges": judge_stats,
            "system_time": "2024-09-24 12:00:00",  # 可以替换为实际系统时间
            "status": "running"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取系统状态失败: {str(e)}")

@router.post("/backup/database")
async def backup_database():
    """备份数据库"""
    try:
        import shutil
        from datetime import datetime
        
        # 创建备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_filename = f"database_backup_{timestamp}.db"
        backup_path = f"data/exports/{backup_filename}"
        
        # 复制数据库文件
        shutil.copy2("data/database.db", backup_path)
        
        return {
            "message": "数据库备份成功",
            "backup_file": backup_filename,
            "backup_path": backup_path
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"备份失败: {str(e)}")